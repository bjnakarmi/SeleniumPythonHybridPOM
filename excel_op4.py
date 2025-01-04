# Merging cells together, write data in the excel spreadsheet and style them accordingly.

from openpyxl import Workbook, load_workbook

wb = load_workbook('Test.xlsx')
ws = wb.active
# ws.merge_cells('A1:D1')       #merging multiple cells
# ws.unmerge_cells('A1:D1')     #unmerging cells
# wb.save('Test.xlsx')


# Inserting and Deleting rows and cols.

# ws.insert_rows(3)
# ws.delete_rows(3)
# ws.insert_cols(3)
# ws.delete_cols(3)
# wb.save('Test.xlsx')

# Copy and move portions of the worksheet.

ws.move_range("C3:D5", -2, 2)           # we use '-' to go cells before or upward the current row/col and vice versa
wb.save('Test.xlsx')

