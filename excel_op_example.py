# Practical example, formulas and styling

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = {
    "employee1": {
        "name":       "sonoo1",
        "salary":      560001,
        "married":    'true'
    },
    "employee2": {
        "name":       "sonoo2",
        "salary":      560002,
        "married":    'true'
    },
    "employee3": {
        "name":       "sonoo3",
        "salary":      560003,
        "married":    'false'
    }
}

wb = Workbook()
ws = wb.active
ws.title = 'Data'

headings =list(data['employee1'].keys())
ws.append(headings)
for person in data:
    salary = list(data[person].values())
    ws.append(salary)

wb.save("NewEmp.xlsx")

# Styling Our Cells
for row in range(1,4):
    ws[get_column_letter(row) + '1'].font = Font(bold=True, color='0099CCFF')

wb.save("NewEmp.xlsx")