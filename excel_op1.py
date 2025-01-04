# Creating new worksheet inside an workbook/excel file
from openpyxl import Workbook, load_workbook

wb = load_workbook('Grades.xlsx')
# wb.create_sheet('Sheet4')
# wb.save('Grades.xlsx')
print(wb.sheetnames)