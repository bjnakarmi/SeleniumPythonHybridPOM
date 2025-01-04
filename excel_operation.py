from openpyxl import Workbook, load_workbook

wb = load_workbook('Grades.xlsx')  # wb = workbook
ws = wb.active                     # ws = worksheet
print(ws['A1'].value)               # to look at the value of the worksheet[A1]

# ws['A1'] = 'NAME'                 # no need to use .value . Only when accessing the value
# wb.save('Grades.xlsx')

print(wb.sheetnames)                # list of sheets in the workbook
print(wb.active)
ws = wb['Sheet2']                   # Switching sheets/active sheets
print(ws) 


