# Loop through different rows and columns, access and modify different cell values in a specific range

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('Test.xlsx')
wb.create_sheet('Sheet2')
ws = wb.active

# to get the row/column name
for row in range(1, 11):
    for col in range(1,5):
        char = get_column_letter(col)      # takes an integer from 1 - 26 and characterizes it like A,B etc
        print(ws[char + str(row)].value)


